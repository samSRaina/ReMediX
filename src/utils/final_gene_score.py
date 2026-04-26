from functools import lru_cache
import math
import openpyxl
from pathlib import Path
from ..clients import creeds_client

_EXCLUDED_SHEETS = {"reactome"}

@lru_cache(maxsize=1)
def load_excel_sheets() -> dict[str, list[list]]:
    """Read and cache all supported Excel sheets from the PPInteraction dataset."""
    from datetime import datetime, date

    def _serialise(v):
        if v is None:
            return None
        if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
            return None
        if isinstance(v, (datetime, date)):
            return str(v)
        return v

    data_dir = Path(__file__).resolve().parent.parent / "data" / "PPInteraction" / "xlsxData"
    result: dict[str, list[list]] = {}

    for excel_path in sorted(data_dir.glob("*.xlsx")):
        if excel_path.name.startswith("~$"):
            continue

        wb = openpyxl.load_workbook(excel_path, data_only=True)
        base_name = excel_path.stem.strip()

        for sheet_name in wb.sheetnames:
            if sheet_name.strip().lower() in _EXCLUDED_SHEETS or base_name.lower().find("reactome") >= 0:
                continue

            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                cleaned = [_serialise(c) for c in row]
                # trim trailing None cells to keep payload compact
                while cleaned and cleaned[-1] is None:
                    cleaned.pop()
                rows.append(cleaned)

            if len(wb.sheetnames) == 1:
                key = base_name
            else:
                key = f"{base_name} - {sheet_name.strip()}"
            result[key] = rows

    return result


def calculate_final_score(genes: str, disease: str) -> dict:
    """
    Calculate final score from disease signature scores:
    - numerator: sum of beneficially matched disease-signature gene scores
    - denominator: sum of all disease-signature scores (up + down rows)
    """
    if not genes:
        return {"score": 0.0, "genes_counted": []}
    if not disease or not disease.strip():
        raise ValueError("Disease parameter is required for final score calculation")

    gene_list = [g.strip() for g in genes.split(",") if g.strip()]

    match_data = creeds_client.match_gene_set(gene_list, disease)
    numerator = float(match_data.get("beneficial_sum", match_data.get("beneficial_disease_score_total", 0.0)) or 0.0)
    harmful_sum = float(match_data.get("harmful_sum", 0.0) or 0.0)
    denominator = numerator + harmful_sum
    genes_found = [
        str(item.get("gene", "")).strip().upper()
        for item in (match_data.get("beneficial_disease_genes") or [])
        if str(item.get("gene", "")).strip()
    ]

    final_result = 0.0 if denominator == 0 else numerator / denominator
    rounded_score = round(final_result, 4)

    return {
        "score": rounded_score,
        "numerator": round(numerator, 4),
        "denominator": round(denominator, 4),
        "genes_counted": genes_found,
        "beneficial_sum": round(numerator, 4),
        "harmful_sum": round(harmful_sum, 4),
        "final_score": rounded_score,
        "interpretation": match_data.get("interpretation"),
        "coverage": round(float(match_data.get("coverage", 0.0) or 0.0), 4),
        "matched_gene_count": match_data.get("matched_gene_count", 0),
        "input_gene_count": len(gene_list),
    }
