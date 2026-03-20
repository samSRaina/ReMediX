import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from pathlib import Path


def calculate_gene_scores():
    """Calculate and save gene scores to the Excel file."""
    # Load the Excel file
    excel_file = Path(__file__).parent.parent / "data" / 'data_set.xlsx'

    # Define sheet names and their gene/score columns
    sheets_config = {
        'GeneCards ': {'gene_col': 'Gene Symbol', 'score_col': 'Final Score '},
        'GEO': {'gene_col': 'Gene Symbol', 'score_col': 'Final Score'},
        'Final Reactome Sheet ': {'gene_col': 'Unique Genes', 'score_col': 'Final Score '},
        'OpenTargets': {'gene_col': 'Gene Symbol', 'score_col': 'Final Score '}
    }

    # Dictionary to store accumulated scores
    gene_scores = {}

    # Process each sheet
    for sheet_name, config in sheets_config.items():
        print(f"Processing {sheet_name}...")

        try:
            df = pd.read_excel(excel_file, sheet_name=sheet_name)

            gene_col = config['gene_col']
            score_col = config['score_col']

            # Iterate through rows and accumulate scores
            for _, row in df.iterrows():
                gene = row[gene_col]
                score = row[score_col]

                # Skip NaN values and invalid entries
                if pd.notna(gene) and pd.notna(score):
                    try:
                        score = float(score)
                        # Split genes by /// to treat each as unique
                        genes_list = str(gene).split('///')
                        for g in genes_list:
                            g = g.strip()
                            if g:  # Only add non-empty genes
                                if g not in gene_scores:
                                    gene_scores[g] = 0
                                gene_scores[g] += score
                    except (ValueError, TypeError):
                        pass

        except Exception as e:
            print(f"  Error processing {sheet_name}: {e}")

    # Create a DataFrame from the results
    result_df = pd.DataFrame(
        list(gene_scores.items()),
        columns=['Gene', 'Total Score']
    ).sort_values('Total Score', ascending=False)

    print(f"\n✓ Processed {len(gene_scores)} unique genes")
    print(f"\nTop 10 genes by score:")
    print(result_df.head(10))

    # Add the result to the Excel file using openpyxl (safer)
    try:
        # Load the existing workbook
        wb = load_workbook(excel_file)

        # Remove Gene Scores sheet if it exists
        if 'Gene Scores' in wb.sheetnames:
            del wb['Gene Scores']

        # Create new sheet
        ws = wb.create_sheet('Gene Scores')

        # Write headers
        headers = ['Gene', 'Total Score']
        ws.append(headers)

        # Write data rows
        for gene, score in result_df.values:
            ws.append([gene, score])

        # Save the workbook
        wb.save(excel_file)

        print(f"\n✓ Results saved to 'Gene Scores' sheet in {excel_file}")
    except PermissionError:
        print(f"\nERROR: Cannot write to Excel file - it's currently open in another application.")
        print(f"   Please close the Excel file and try again.")
        print(f"   File location: {excel_file}")
    except Exception as e:
        print(f"\nERROR: Failed to save results: {e}")


if __name__ == "__main__":
    calculate_gene_scores()
