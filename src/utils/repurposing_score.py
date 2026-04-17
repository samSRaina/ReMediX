from __future__ import annotations

from functools import lru_cache
from pathlib import Path
from typing import Any

import openpyxl
import requests

from ..clients import chembl_client, creeds_client, drugbank_client, pubchem_client

_DATA_ROOT = Path(__file__).resolve().parent.parent / "data"
_CONFIG_ROOT = Path(__file__).resolve().parent.parent / "config"

_EXPECTED_LOCAL_FILES = {
    "creeds_disease_signatures": _DATA_ROOT / "CREEDS" / "disease_signatures-v1.0.json",
    "creeds_perturbations": _DATA_ROOT / "CREEDS" / "single_drug_perturbations-v1.0.json",
    "drugbank_xml": _DATA_ROOT / "drugBank" / "full database.xml",
    "main_excel": _DATA_ROOT / "data_set.xlsx",
    "genecards_geo_excel": _DATA_ROOT / "geneCards" / "GEO DATA.xlsx",
}

_DISease_config_path = _CONFIG_ROOT / "disease_class_config.json"
_DISEASE_RULES_PATH = _CONFIG_ROOT / "disease_class_rules.json"


@lru_cache(maxsize=1)
def _load_json_file(path: Path) -> dict:
    if not path.exists():
        return {}
    import json

    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}


@lru_cache(maxsize=1)
def discover_local_data_status() -> dict[str, Any]:
    optional_patterns = ["*reactome*", "*opentarget*", "*geo*", "*disease*"]
    present = {}
    missing = []

    for key, file_path in _EXPECTED_LOCAL_FILES.items():
        exists = file_path.exists()
        present[key] = {
            "path": str(file_path),
            "exists": exists,
            "required": key in {"creeds_disease_signatures", "creeds_perturbations"},
        }
        if not exists:
            missing.append(key)

    optional_hits: list[str] = []
    if _DATA_ROOT.exists():
        all_files = [p for p in _DATA_ROOT.rglob("*") if p.is_file()]
        for p in all_files:
            lower = p.name.lower()
            if any(pattern.replace("*", "") in lower for pattern in optional_patterns):
                optional_hits.append(str(p))

    return {
        "data_root": str(_DATA_ROOT),
        "expected": present,
        "missing": missing,
        "optional_hits": sorted(set(optional_hits)),
    }


def _parse_genes(genes: str) -> list[str]:
    seen: set[str] = set()
    parsed: list[str] = []
    for raw in (genes or "").split(","):
        gene = raw.strip().upper()
        if gene and gene not in seen:
            seen.add(gene)
            parsed.append(gene)
    return parsed


def _clamp(value: float, lo: float = 0.0, hi: float = 1.0) -> float:
    return max(lo, min(hi, float(value)))


def _load_disease_class_config() -> dict[str, dict]:
    defaults = {
        "default": {
            "weights": {"alignment": 0.4, "pathway": 0.35, "evidence": 0.25},
            "safety_tolerance": 0.5,
            "scale": 1.0,
        }
    }
    loaded = _load_json_file(_DISease_config_path)
    if not loaded:
        return defaults
    if "default" not in loaded:
        loaded["default"] = defaults["default"]
    return loaded


def _load_disease_class_rules() -> dict[str, list[str]]:
    loaded = _load_json_file(_DISEASE_RULES_PATH)
    if not loaded:
        return {}
    return {k: [str(x).lower() for x in (v or [])] for k, v in loaded.items()}


def map_disease_class(disease: str) -> str:
    text = (disease or "").strip().lower()
    if not text:
        return "default"

    rules = _load_disease_class_rules()
    for disease_class, keywords in rules.items():
        if any(keyword in text for keyword in keywords):
            return disease_class
    return "default"


def _get_class_profile(disease_class: str) -> dict[str, Any]:
    cfg = _load_disease_class_config()
    profile = cfg.get(disease_class) or cfg.get("default") or {}
    weights = profile.get("weights", {})

    alignment = float(weights.get("alignment", 0.4) or 0.4)
    pathway = float(weights.get("pathway", 0.35) or 0.35)
    evidence = float(weights.get("evidence", 0.25) or 0.25)
    total = alignment + pathway + evidence
    if total <= 0:
        alignment, pathway, evidence = 0.4, 0.35, 0.25
        total = 1.0

    return {
        "weights": {
            "alignment": alignment / total,
            "pathway": pathway / total,
            "evidence": evidence / total,
        },
        "safety_tolerance": _clamp(float(profile.get("safety_tolerance", 0.5) or 0.5), 0.0, 1.0),
        "scale": max(0.0, float(profile.get("scale", 1.0) or 1.0)),
    }


def _resolve_compound_identity(compound_name: str | None, inchikey: str | None) -> tuple[str | None, dict[str, Any]]:
    used = {"pubchem": False}
    resolved_inchikey = (inchikey or "").strip() or None

    if not resolved_inchikey and compound_name:
        used["pubchem"] = True
        resolved_inchikey = pubchem_client.PubChemClient().get_inchikey(compound_name)

    return resolved_inchikey, used


def _compute_alignment_layer(genes: list[str], disease: str) -> dict[str, Any]:
    match = creeds_client.match_gene_set(genes, disease)
    beneficial_sum = float(match.get("beneficial_sum", 0.0) or 0.0)
    harmful_sum = float(match.get("harmful_sum", 0.0) or 0.0)
    denom = beneficial_sum + harmful_sum
    alignment_score = 0.0 if denom == 0 else beneficial_sum / denom

    beneficial_targets = [row.get("gene") for row in (match.get("up_genes") or []) if row.get("gene")]
    harmful_targets = [row.get("gene") for row in (match.get("down_genes") or []) if row.get("gene")]

    return {
        "alignment_score": round(_clamp(alignment_score), 4),
        "beneficial_sum": round(beneficial_sum, 4),
        "harmful_sum": round(harmful_sum, 4),
        "beneficial_targets": beneficial_targets,
        "harmful_targets": harmful_targets,
        "match_payload": match,
    }


def _pathway_rows_from_local_excel() -> list[dict[str, Any]]:
    excel_path = _EXPECTED_LOCAL_FILES["main_excel"]
    if not excel_path.exists():
        return []

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    rows: list[dict[str, Any]] = []

    for sheet_name in wb.sheetnames:
        if "reactome" not in sheet_name.lower() and "target" not in sheet_name.lower():
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            compact = [str(cell).strip() for cell in row if cell not in (None, "")]
            if compact:
                rows.append({"sheet": sheet_name, "cells": compact})

    return rows


def _compute_pathway_layer(genes: list[str]) -> dict[str, Any]:
    rows = _pathway_rows_from_local_excel()
    if not rows:
        return {
            "pathway_coherence_score": 0.0,
            "pathways_hit": [],
            "pathway_hit_count": 0,
            "total_pathway_rows": 0,
            "used_local_files": False,
        }

    genes_set = set(genes)
    pathway_hits = []
    for row in rows:
        cells_upper = {c.upper() for c in row["cells"]}
        overlap = sorted(list(cells_upper & genes_set))
        if overlap:
            pathway_hits.append(
                {
                    "source": row["sheet"],
                    "genes": overlap,
                    "preview": row["cells"][:5],
                }
            )

    unique_pathways = {(entry["source"], tuple(entry["genes"])) for entry in pathway_hits}
    score = 0.0 if len(rows) == 0 else len(unique_pathways) / len(rows)

    return {
        "pathway_coherence_score": round(_clamp(score), 4),
        "pathways_hit": pathway_hits[:25],
        "pathway_hit_count": len(unique_pathways),
        "total_pathway_rows": len(rows),
        "used_local_files": True,
    }


def _fetch_clinical_trials_count(disease: str) -> int | None:
    if not disease:
        return None

    url = "https://clinicaltrials.gov/api/query/study_fields"
    params = {
        "expr": disease,
        "fields": "NCTId",
        "min_rnk": 1,
        "max_rnk": 1,
        "fmt": "json",
    }

    try:
        response = requests.get(url, params=params, timeout=8)
        response.raise_for_status()
        payload = response.json()
        return int(payload.get("StudyFieldsResponse", {}).get("NStudiesFound", 0) or 0)
    except Exception:
        return None


def _local_evidence_hits(disease: str, genes: list[str]) -> int:
    excel_path = _EXPECTED_LOCAL_FILES["main_excel"]
    if not excel_path.exists():
        return 0

    disease_lower = (disease or "").strip().lower()
    genes_set = {g.upper() for g in genes}
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    hits = 0

    for sheet_name in wb.sheetnames:
        lowered = sheet_name.lower()
        if all(k not in lowered for k in ("target", "evidence", "reactome", "geo")):
            continue

        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            values = [str(v).strip() for v in row if v not in (None, "")]
            if not values:
                continue
            text = " ".join(values).lower()
            row_genes = {value.upper() for value in values}
            if (disease_lower and disease_lower in text) or (row_genes & genes_set):
                hits += 1

    return hits


def _compute_evidence_layer(disease: str, genes: list[str]) -> dict[str, Any]:
    local_hits = _local_evidence_hits(disease, genes)
    trials_count = _fetch_clinical_trials_count(disease)

    local_component = min(1.0, local_hits / 250.0)
    trial_component = 0.0 if trials_count is None else min(1.0, trials_count / 2000.0)

    score = 0.7 * local_component + 0.3 * trial_component

    return {
        "evidence_score": round(_clamp(score), 4),
        "local_evidence_hits": local_hits,
        "clinical_trials_count": trials_count,
        "evidence_summary": {
            "local_evidence_strength": round(local_component, 4),
            "clinical_trials_strength": round(trial_component, 4),
        },
    }


def _fetch_openfda_warnings(compound_name: str) -> dict[str, Any]:
    if not compound_name:
        return {"available": False, "boxed_warning": False, "warnings_present": False}

    url = "https://api.fda.gov/drug/label.json"
    params = {"search": f'openfda.generic_name:"{compound_name}"', "limit": 1}

    try:
        response = requests.get(url, params=params, timeout=8)
        response.raise_for_status()
        result = (response.json().get("results") or [{}])[0]
        boxed = bool(result.get("boxed_warning"))
        warnings = bool(result.get("warnings"))
        return {
            "available": True,
            "boxed_warning": boxed,
            "warnings_present": warnings,
        }
    except Exception:
        return {"available": False, "boxed_warning": False, "warnings_present": False}


def _compute_safety_layer(compound_name: str | None, inchikey: str | None, safety_tolerance: float) -> dict[str, Any]:
    risk = 0.0
    notes: list[str] = []

    resolved_name = (compound_name or "").strip()
    drugbank_data = None
    if inchikey:
        drugbank_data = drugbank_client.DrugBankClient().search_drug_by_inchikey(inchikey)

    if drugbank_data:
        groups = [str(g).lower() for g in (drugbank_data.get("groups") or [])]
        if any(g in {"withdrawn", "illicit"} for g in groups):
            risk += 0.4
            notes.append("DrugBank group indicates elevated risk")

    openfda = _fetch_openfda_warnings(resolved_name)
    if openfda.get("boxed_warning"):
        risk += 0.35
        notes.append("openFDA boxed warning found")
    elif openfda.get("warnings_present"):
        risk += 0.2
        notes.append("openFDA warnings section present")

    tolerated_risk = max(0.0, risk - (0.2 * safety_tolerance))
    safety_modifier = _clamp(1.0 - tolerated_risk, 0.2, 1.0)

    return {
        "safety_modifier": round(safety_modifier, 4),
        "raw_risk": round(risk, 4),
        "safety_notes": notes,
        "openfda": openfda,
        "drugbank_found": bool(drugbank_data),
    }


def _chembl_context(inchikey: str | None) -> dict[str, Any]:
    if not inchikey:
        return {"used": False, "target_count": 0, "targets": []}

    client = chembl_client.ChEMBLClient()
    activities = client.get_by_inchikey(inchikey)
    targets = sorted({a.get("gene_symbol") for a in activities if a.get("gene_symbol") and a.get("gene_symbol") != "--"})
    return {
        "used": True,
        "target_count": len(targets),
        "targets": targets,
    }


def calculate_repurposing_score(genes: str, disease: str, compound_name: str | None = None, inchikey: str | None = None) -> dict[str, Any]:
    if not disease or not disease.strip():
        raise ValueError("Disease parameter is required")

    gene_list = _parse_genes(genes)
    if not gene_list:
        raise ValueError("No genes provided")

    local_data_status = discover_local_data_status()
    disease_class = map_disease_class(disease)
    profile = _get_class_profile(disease_class)
    weights = profile["weights"]

    resolved_inchikey, identity_sources = _resolve_compound_identity(compound_name, inchikey)
    chembl = _chembl_context(resolved_inchikey)

    alignment = _compute_alignment_layer(gene_list, disease)
    pathway = _compute_pathway_layer(gene_list)
    evidence = _compute_evidence_layer(disease, gene_list)
    safety = _compute_safety_layer(compound_name, resolved_inchikey, profile["safety_tolerance"])

    base_score = (
        (weights["alignment"] * alignment["alignment_score"])
        + (weights["pathway"] * pathway["pathway_coherence_score"])
        + (weights["evidence"] * evidence["evidence_score"])
    )
    final_score = _clamp(base_score * safety["safety_modifier"] * profile["scale"])

    missing_components = []
    if local_data_status["missing"]:
        missing_components.extend(local_data_status["missing"])
    if evidence["clinical_trials_count"] is None:
        missing_components.append("clinical_trials_api")
    if not safety["openfda"].get("available"):
        missing_components.append("openfda_api")

    provenance = {
        "creeds": {"used": True, "dataset": str(_EXPECTED_LOCAL_FILES["creeds_disease_signatures"])},
        "chembl": chembl,
        "drugbank": {"used": bool(safety["drugbank_found"]), "dataset": str(_EXPECTED_LOCAL_FILES["drugbank_xml"])},
        "pubchem": identity_sources,
        "clinical_trials": {"used": evidence["clinical_trials_count"] is not None},
        "openfda": {"used": safety["openfda"].get("available", False)},
        "local_files": local_data_status,
    }

    return {
        "disease": disease,
        "disease_class": disease_class,
        "inchikey": resolved_inchikey,
        "alignment_score": alignment["alignment_score"],
        "pathway_coherence_score": pathway["pathway_coherence_score"],
        "evidence_score": evidence["evidence_score"],
        "safety_modifier": safety["safety_modifier"],
        "repurposing_score": round(final_score, 4),
        "weights": {k: round(v, 4) for k, v in weights.items()},
        "explanations": {
            "beneficial_targets": alignment["beneficial_targets"][:50],
            "harmful_targets": alignment["harmful_targets"][:50],
            "pathways_hit": pathway["pathways_hit"],
            "evidence_summary": evidence["evidence_summary"],
            "safety_summary": {
                "notes": safety["safety_notes"],
                "raw_risk": safety["raw_risk"],
            },
        },
        "data_provenance": provenance,
        "missing_components": sorted(set(missing_components)),
    }
