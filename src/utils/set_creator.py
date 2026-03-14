from pathlib import Path

import openpyxl
from collections import Counter

def extract_unique_genes(file_path):

    wb = openpyxl.load_workbook(file_path)
    ws = wb['Top Reactome']

    genes = []
    for row in range(2, 47):
        for col in range(3, 454):  # C to QK
            cell = ws.cell(row, col)
            if cell.value is not None and str(cell.value).strip():
                genes.append(str(cell.value).strip())

    # Count unique genes
    gene_counts = Counter(genes)
    unique_genes = sorted(gene_counts.keys())

    # Add headers at row 48
    ws.cell(48, 1, "Unique Genes")
    ws.cell(48, 2, "Count")

    # Add unique genes and counts starting at row 49
    for idx, gene in enumerate(unique_genes, start=49):
        ws.cell(idx, 1, gene)
        ws.cell(idx, 2, gene_counts[gene])

    wb.save(file_path)

    print(f"✅ SUCCESS!")
    print(f"✅ Added {len(unique_genes)} unique genes with counts")
    print(f"✅ Results start at row 48")
    print(f"✅ Original data (rows 1-46) completely untouched")


if __name__ == "__main__":
    filepath = Path(__file__).parent.parent/"data"/"data_set.xlsx"
    extract_unique_genes(filepath)