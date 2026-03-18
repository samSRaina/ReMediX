from fastapi import APIRouter, HTTPException
from typing import Optional
from ..clients import pubchem_client, drugbank_client, chembl_client, creeds_client, geneCards_client
import openpyxl
from pathlib import Path

router= APIRouter(prefix="/api")

# PubChem database endpoints
@router.get("/compound/name/{name}/properties")
async def get_properties_by_name_api(name: str):
    return pubchem_client.PubChemClient().search_by_name(name)


@router.get("/compound/smile/{smile}/properties")
async def get_properties_by_smile_api(smile: str):
    return pubchem_client.PubChemClient().search_by_smile(smile)


# DrugBank database endpoints
@router.get("/drugbank/inchikey/{inchikey}/properties")
async def get_properties_by_inchikey(inchikey: str):
    result = drugbank_client.DrugBankClient()
    if result is None:
        raise HTTPException(status_code=404, detail=f"Drug '{inchikey}' not found in DrugBank database")
    return result.search_drug_by_inchikey(inchikey)


# ChEMBL database endpoints
@router.get("/chembl/inchikey/{inchikey}/bioactivity")
async def get_bioactivity_by_inchikey(inchikey: str, standard_type: Optional[str] = None):
    chembl = chembl_client.ChEMBLClient()
    result = chembl.get_by_inchikey(inchikey, standard_type)
    if not result:
        raise HTTPException(status_code=404, detail=f"No bioactivity data found for '{inchikey}'")
    gene_set = chembl.get_gene_set(inchikey)
    return {"activities": result, "gene_set": sorted(gene_set)}

# CREEDS match endpoint — matches each gene against disease signatures (pulmonary hypertension)
@router.get("/match")
async def get_gene_match(genes: str):
    """genes = comma-separated gene symbols"""
    gene_list = [g.strip() for g in genes.split(",") if g.strip()]
    if not gene_list:
        raise HTTPException(status_code=400, detail="No genes provided")
    return creeds_client.match_gene_set(gene_list)

@router.get("/chembl/inchikey/{inchkey}/bioactivity/{target_chembl_id}/target")
async def get_target_data(target_chembl_id: str):
    result = chembl_client.ChEMBLClient().get_target_data(target_chembl_id)
    if not result:
        raise HTTPException(status_code =404, detail=f"No target data found for {target_chembl_id}")
    return result

@router.get("/geneAnalysis/accession/{accession_id}")
async def get_gene_analysis(accession_id: str, disease: str):
    disease_signatures = creeds_client.get_disease_signatures(disease)
    accession_object = creeds_client.CreedsClient(accession_id)
    single_gene_perturbations = accession_object.get_single_gene_perturbations()
    return accession_object.match_genes(disease_signatures, single_gene_perturbations )

@router.get("/geneExpressions")
async def get_gene_expressions(page: int = 1, page_size: int = 50, search: Optional[str] = None):
    return geneCards_client.get_geo_data(page, page_size, search)


@router.get("/excelData/meta")
async def get_excel_meta():
    """Return sheet names, column headers and row counts (lightweight)."""
    sheets = _load_excel_sheets()
    meta = {}
    for name, rows in sheets.items():
        headers = rows[0] if rows else []
        meta[name] = {
            "headers": headers,
            "totalRows": max(0, len(rows) - 1),   # exclude header
        }
    return {"sheetNames": list(sheets.keys()), "meta": meta}


@router.get("/excelData/sheet")
async def get_excel_sheet(name: str, page: int = 1, page_size: int = 100):
    """Return a paginated slice of one sheet's data rows."""
    sheets = _load_excel_sheets()
    if name not in sheets:
        raise HTTPException(status_code=404, detail=f"Sheet '{name}' not found")

    all_rows = sheets[name]
    headers = all_rows[0] if all_rows else []
    data_rows = all_rows[1:]                      # everything after header

    total = len(data_rows)
    total_pages = max(1, (total + page_size - 1) // page_size)
    page = max(1, min(page, total_pages))
    start = (page - 1) * page_size
    end = start + page_size

    return {
        "headers": headers,
        "data": data_rows[start:end],
        "page": page,
        "pageSize": page_size,
        "total": total,
        "totalPages": total_pages,
    }


@router.get("/diseaseSignature/table")
async def get_disease_signature_table(disease: str = "pulmonary hypertension", page: int = 1, page_size: int = 100):
    """Export one disease signature to JSON and return a paginated table payload."""
    try:
        payload = creeds_client.export_disease_signature_table(disease)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc

    rows = payload.get("rows", [])
    headers = payload.get("headers", [])
    total = len(rows)
    total_pages = max(1, (total + page_size - 1) // page_size)
    page = max(1, min(page, total_pages))
    start = (page - 1) * page_size
    end = start + page_size

    return {
        "disease": payload.get("disease", disease),
        "headers": headers,
        "data": rows[start:end],
        "page": page,
        "pageSize": page_size,
        "total": total,
        "totalPages": total_pages,
    }


# ── helpers ────────────────────────────────────────────────────
from functools import lru_cache
import math

_EXCLUDED_SHEETS = ["Reactome"]

@lru_cache(maxsize=1)
def _load_excel_sheets() -> dict[str, list[list]]:
    """Read all non-excluded sheets once and cache them."""
    from datetime import datetime, date

    def _serialise(v):
        if v is None:
            return None
        if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
            return None
        if isinstance(v, (datetime, date)):
            return str(v)
        return v

    excel_path = Path(__file__).resolve().parent.parent / "data" / "data_set.xlsx"
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    result: dict[str, list[list]] = {}

    for sheet_name in wb.sheetnames:
        if sheet_name in _EXCLUDED_SHEETS:
            continue
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            cleaned = [_serialise(c) for c in row]
            # trim trailing None cells (Top Reactome has 453 cols, mostly empty)
            while cleaned and cleaned[-1] is None:
                cleaned.pop()
            rows.append(cleaned)
        result[sheet_name] = rows

    return result
